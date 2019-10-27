# coding:utf-8

import openpyxl
import pandas as pd
import copy


def openpyxl_to_pandas(data, sheet_index=None, header=None):
    """
    convert data from openpyxl to DataFrame in Pandas
    usage:
        load_workbook
        convert
        pd.
        pd.
        ...
    :param data: data from openpyxl, including workbook, worksheet and cell
        Workbook: return active sheet value
        Worksheet: return sheet value
        Cell: return cell value
    :param sheet_index: int, None; which sheet in workbook
    :param header: int, None, default None
        Row (0-indexed) to use for the column labels of the parsed DataFrame.
        If a list of integers is passed those row positions will be combined
        into a MultiIndex. Use None if there is no header.
    :return: DataFrame
    """

    _df = None
    title = None

    # deal workbook
    if isinstance(data, openpyxl.workbook.workbook.Workbook):
        if sheet_index is None:
            # active sheet not only 1st sheet
            _df = pd.DataFrame(data.active.values)
            title = [c.value for c in data.active[1]]
        if isinstance(sheet_index, int):
            name = data.get_sheet_names()[sheet_index]
            sh = data.get_sheet_by_name(name)
            _df = pd.DataFrame(sh.values)
            title = [c.value for c in sh[1]]

    # deal worksheet
    if isinstance(data, openpyxl.workbook.workbook.Worksheet):
        # 1.generate dataframe
        tmp = data.values
        _df = pd.DataFrame(tmp)
        title = [c.value for c in data[1]]

    # deal cell
    if isinstance(data, tuple):
        tmp = []
        for row in data:
            tmp.append([c.value for c in row])
        _df = pd.DataFrame(tmp)
        title = [c.value for c in data[0]]

    # deal header
    if header is None:
        return df
    if header == 0:
        _df = _df.drop([0], axis=0)
        _df.columns = title
        _df = _df.reset_index(drop=True)
        return df
    if header > 0 and isinstance(header, int):
        title = [c.value for c in data[header + 1]]
        _df = _df.drop([i for i in range(0, header)], axis=0)
        _df.columns = title
        _df = _df.reset_index(drop=True)
        return _df


def pandas_to_openpyxl(df, whichExcel=None, whichSheet=None,
                               startRow=1, startCol=1,
                               index=False, header=True):
    """
    convert pandas data to openpyxl and write to excel
    usage:
        load_workbook
        convert(df, wb, 'sheet')
        convert(df, whichSheet=ws)
        ...
        save
    :param df: raw data
    :param whichExcel:
        openpyxl.workbook.workbook.Workbook or None
        target excel
    :param whichSheet:
        openpyxl.workbook.workbook.Worksheet or str
        target sheet
    :param startRow: start with 1
    :param startCol: start with 1 or A
    :param index: true means save pandas' dataframe index
    :param header: true means save pandas' dataframe header
    :return:
    """

    # 1.get target sheet
    _sh = None
    if whichExcel is None:
        if isinstance(whichSheet, openpyxl.workbook.workbook.Worksheet):
            _sh = whichSheet
        else:
            return 'ERROR: please input openpyxl.workbook.workbook.Worksheet'
    elif isinstance(whichExcel, openpyxl.workbook.workbook.Workbook):
        if isinstance(whichSheet, str):
            _sh = whichExcel[whichSheet]
        else:
            return 'ERROR: please input str sheet name'
    else:
        return 'ERROR: please input openpyxl.workbook.workbook.Workbook'

    # 2.reset row and col
    srow = startRow
    scol = None
    if isinstance(startCol, int):
        scol = startCol
    elif isinstance(startCol, str):
        from openpyxl.utils import column_index_from_string
        scol = column_index_from_string(startCol)

    # 3.backup df
    data = df.copy()

    # 4.write index and header
    if index and header:
        ind = data.index
        hea = data.columns
        row = srow + 1
        for i in ind:
            _sh.cell(row=row, column=scol, value=i)
            row = row + 1
        col = scol + 1
        for h in hea:
            _sh.cell(row=srow, column=col, value=h).style='Pandas'
            col = col + 1
        srow = srow + 1
        scol = scol + 1
    elif index:
        ind = data.index
        row = copy.copy(srow)
        for i in ind:
            _sh.cell(row=row, column=scol, value=i)
            row = row + 1
        scol = scol + 1
    elif header:
        hea = data.columns
        col = copy.copy(scol)
        for h in hea:
            _sh.cell(row=srow, column=col, value=h).style='Pandas'
            col = col + 1
        srow = srow + 1

    # 5.write data to excel
    tmp = data.values.tolist()
    for r in tmp:
        # write data
        col = copy.copy(scol)
        for c in r:
            _sh.cell(row=srow, column=col, value=c)
            col = col + 1
        srow = srow + 1
    return 'success'


def clear_sheet(whichExcel=None, whichSheet=None, data_only=True):
    _sh = None
    if whichExcel is None:
        if isinstance(whichSheet, openpyxl.workbook.workbook.Worksheet):
            _sh = whichSheet
        else:
            return 'ERROR: please input openpyxl.workbook.workbook.Worksheet'
    elif isinstance(whichExcel, openpyxl.workbook.workbook.Workbook):
        if isinstance(whichSheet, str):
            _sh = whichExcel[whichSheet]
        else:
            return 'ERROR: please input str sheet name'
    else:
        return 'ERROR: please input openpyxl.workbook.workbook.Workbook'

    from openpyxl.styles import Border
    if data_only:
        for r in _sh.iter_rows():
            for c in r:
                c.border = Border() # delete border
                c.value = ''


if __name__ == '__main__':
    wb = openpyxl.load_workbook('./result.xlsx')
    sheet = wb['区县汇总']
    clear_sheet(whichSheet=sheet)
    # cell = sheet['b2:F6']  # 不区分大小写
    #w = convert_openpyxl_to_pandas(wb)
    #s = convert_openpyxl_to_pandas(sheet, 0)
    #c = convert_openpyxl_to_pandas(cell, 0)

    import numpy as np
    dates = pd.date_range('20130101', periods=6)
    df = pd.DataFrame(
        np.arange(24).reshape(
            (6, 4)), index=dates, columns=[
                'A', 'B', 'C', 'D'])
    print(df)
    pandas_to_openpyxl(df, wb, '区县汇总',
                               index=True, header=True,
                               startRow=1, startCol=5)

    pandas_to_openpyxl(df, wb, '区县汇总',
                               index=False, header=True,
                               startRow=10, startCol=5)

    pandas_to_openpyxl(df, wb, '区县汇总',
                               index=True, header=False,
                               startRow=19, startCol=5)
    wb.save('./result.xlsx')
